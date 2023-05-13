import openpyxl
from random import randint as rand

workbook = openpyxl.Workbook()

worksheet = workbook.active

data = []
features = {'A':[0,0,0,0,0,0,0,0,0],'E':[0,0,0,0],'C':[0,0,0,0,0,0]}
for i in range(1000):
    for j in range(9):
     features['A'][j] = rand(0,10);  #0--->9 
    for j in range(4):
     features['E'][j] = rand(1,5); # 1,2,3,4   
    for j in range(6):
     features['C'][j] = rand(0,5);   # 0,1,2,3,4,5
    data.append((i+1,features['A'][0],features['A'][1],features['A'][2],features['A'][3],features['A'][4],features['A'][5],features['A'][6],features['A'][7],features['A'][8],features['E'][0],features['E'][1],features['E'][2],features['E'][3],features['C'][0],features['C'][1],features['C'][2],features['C'][3],features['C'][4],features['C'][5],rand(0,6))) 


# loop through the data and write each row to the worksheet
for row_index, row_data in enumerate(data, start=2):
    for i in range(20):
        worksheet.cell(row=row_index, column=i+1, value=row_data[i])

# data = [
#     ('John Doe', 35, 'john.doe@example.com'),
#     ('Jane Smith', 27, 'jane.smith@example.com'),
#     ('Bob Johnson', 42, 'bob.johnson@example.com')
# ]

# # loop through the data and write each row to the worksheet
# for row_index, row_data in enumerate(data, start=2):
#     worksheet.cell(row=row_index, column=1, value=row_data[0])
#     worksheet.cell(row=row_index, column=2, value=row_data[1])
#     worksheet.cell(row=row_index, column=3, value=row_data[2])

# save the workbook
workbook.save('test.xlsx')