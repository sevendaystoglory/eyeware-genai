import openpyxl
import random
from openpyxl import load_workbook

for i in range(50): #creating sheets
    wb2 = load_workbook('datageneration2.xlsx')
    sheetname="model"
    wb2.create_sheet(sheetname)
wb2.save('datageneration2.xlsx')